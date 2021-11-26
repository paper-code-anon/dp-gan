from openpyxl import load_workbook
import matplotlib.pyplot as plt
import os
import numpy as np
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn import linear_model
from sklearn import metrics
import os

# Inputs
folder_name = 'inputs_october_v2'
week_names = [1, 2, 3]

O_S_name_mapping = {
    1: [f'B{i}' for i in range(2, 236)],
    2: [f'C{i}' for i in range(2, 236)],
    3: [f'D{i}' for i in range(2, 236)]
}
S_O_name_mapping = {
    1: [f'C{i}' for i in range(2, 236)],
    2: [f'D{i}' for i in range(2, 236)],
    3: [f'E{i}' for i in range(2, 236)]
}
model_cell_names = [f'A{i}' for i in range(2, 236)]
s_s_cell_names = [f'B{i}' for i in range(2, 236)]
model_names = ['mlp_dp_sgd_64', 'lstm_dp_sgd', 'mlp_dp_loss_64',  'lstm_dp_loss']
# model_names = ['mlp_dp_sgd', 'lstm_dp_sgd', 'mlp_dp_loss', 'lstm_dp_loss']
sheet_names = ['GLM_synthetic']

original_sheet_name = 'GLM_original'
original_result_cell = 'J2'
original_filename = 'Sept_train_on_original_test_on_synthetic.xlsx'

eps_names = ['2_3_4_1']

for week_name in week_names:
    model_result_dict = {model_name: {} for model_name in model_names}
    O_S_cell_names = O_S_name_mapping[week_name]
    S_O_cell_names = S_O_name_mapping[week_name]

    for eps_name in eps_names:
        """Get O->S Results"""
        filename = 'Sept_train_on_original_test_on_synthetic.xlsx'
        wb = load_workbook(folder_name + '/' + filename)
        working_name = filename.strip('.xlsx')

        mse_dict = {}
        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            for model_cell, result_cell in zip(model_cell_names, O_S_cell_names):
                model_name = ws[model_cell].value
                # Only process epsilon breakdowns we are considering
                if model_name.endswith(eps_name):
                    result = ws[result_cell].value
                    mse_dict[model_name + '_' + sheet_name] = result

        model_result_dict['lstm_dp_loss']['O->S'] = [float(val) for key, val in mse_dict.items() if key.startswith('lstm_dp_loss')]
        model_result_dict['lstm_dp_sgd']['O->S'] = [float(val) for key, val in mse_dict.items() if key.startswith('lstm_dp_sgd')]
        # model_result_dict['mlp_dp_loss']['O->S'] = [float(val) for key, val in mse_dict.items() if key.startswith('mlp_dp_loss__')]
        model_result_dict['mlp_dp_loss_64']['O->S'] = [float(val) for key, val in mse_dict.items() if key.startswith('mlp_dp_loss_64')]
        # model_result_dict['mlp_dp_sgd']['O->S'] = [float(val) for key, val in mse_dict.items() if key.startswith('mlp_dp_sgd__')]
        model_result_dict['mlp_dp_sgd_64']['O->S'] = [float(val) for key, val in mse_dict.items() if key.startswith('mlp_dp_sgd_64')]

        """Get S->O Results"""
        filename = 'Sept_train_on_synthetic_test_on_original.xlsx'
        wb = load_workbook(folder_name + '/' + filename)
        working_name = filename.strip('.xlsx')

        mse_dict = {}
        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            for model_cell, result_cell in zip(model_cell_names, S_O_cell_names):
                model_name = ws[model_cell].value
                # Only process epsilon breakdowns we are considering
                if model_name.endswith(eps_name):
                    result = ws[result_cell].value
                    mse_dict[model_name + '_' + sheet_name] = result

        model_result_dict['lstm_dp_loss']['S->O'] = [float(val) for key, val in mse_dict.items() if key.startswith('lstm_dp_loss')]
        model_result_dict['lstm_dp_sgd']['S->O'] = [float(val) for key, val in mse_dict.items() if key.startswith('lstm_dp_sgd')]
        # model_result_dict['mlp_dp_loss']['S->O'] = [float(val) for key, val in mse_dict.items() if key.startswith('mlp_dp_loss__')]
        model_result_dict['mlp_dp_loss_64']['S->O'] = [float(val) for key, val in mse_dict.items() if key.startswith('mlp_dp_loss_64')]
        # model_result_dict['mlp_dp_sgd']['S->O'] = [float(val) for key, val in mse_dict.items() if key.startswith('mlp_dp_sgd__')]
        model_result_dict['mlp_dp_sgd_64']['S->O'] = [float(val) for key, val in mse_dict.items() if key.startswith('mlp_dp_sgd_64')]

        """Get S->S Results"""
        filename = 'Sept_train_on_synthetic_test_on_original.xlsx'
        wb = load_workbook(folder_name + '/' + filename)
        working_name = filename.strip('.xlsx')

        mse_dict = {}
        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            for model_cell, result_cell in zip(model_cell_names, s_s_cell_names):
                model_name = ws[model_cell].value
                # Only process epsilon breakdowns we are considering
                if model_name.endswith(eps_name):
                    result = ws[result_cell].value
                    mse_dict[model_name + '_' + sheet_name] = result

        model_result_dict['lstm_dp_loss']['S->S'] = [float(val) for key, val in mse_dict.items() if key.startswith('lstm_dp_loss')]
        model_result_dict['lstm_dp_sgd']['S->S'] = [float(val) for key, val in mse_dict.items() if key.startswith('lstm_dp_sgd')]
        # model_result_dict['mlp_dp_loss']['S->S'] = [float(val) for key, val in mse_dict.items() if key.startswith('mlp_dp_loss__')]
        model_result_dict['mlp_dp_loss_64']['S->S'] = [float(val) for key, val in mse_dict.items() if key.startswith('mlp_dp_loss_64')]
        # model_result_dict['mlp_dp_sgd']['S->S'] = [float(val) for key, val in mse_dict.items() if key.startswith('mlp_dp_sgd__')]
        model_result_dict['mlp_dp_sgd_64']['S->S'] = [float(val) for key, val in mse_dict.items() if key.startswith('mlp_dp_sgd_64')]

        """Manage original data"""
        wb = load_workbook(folder_name + '/' + original_filename)
        ws = wb[original_sheet_name]
        MSE_test_original = ws[original_result_cell].value

        """Combine results"""
        chart_dict = {}
        for network_type, mse_dict in model_result_dict.items():
            results_csv = pd.DataFrame(mse_dict)
            results_csv.to_csv(network_type + '_results.csv', index=False)
            chart_dict[network_type] = results_csv

        """Plot"""
        width = 1 / (len(model_names))  # the width of the bars
        utility_comparison_4cases_figure1 = plt.figure(figsize=(16, 12))
        # plt.rc('font', size=16)
        plt.rc('legend', fontsize=18)
        plt.rc('axes', titlesize=18)
        plt.rc('axes', labelsize=18)
        plt.rc('xtick', labelsize=18)
        plt.rc('ytick', labelsize=18)
        ax = utility_comparison_4cases_figure1.add_subplot(111)
        error_bar_capsize = 6
        alpha_value = 0.7

        # Original Results
        ax.bar(width,
               MSE_test_original, width * 3,
               color='r', alpha=alpha_value)

        for setting_idx, syn_data_prefix in enumerate(model_names):
            sub_combined_utility_results = chart_dict[syn_data_prefix]
            print(syn_data_prefix)
            print(sub_combined_utility_results)

            ax.bar(setting_idx + 1 + width * 0,
                   np.mean(sub_combined_utility_results['O->S']), width,
                   yerr=np.std(sub_combined_utility_results['O->S']),
                   color='g', alpha=alpha_value, capsize=error_bar_capsize)
            ax.bar(setting_idx + 1 + width * 1,
                   np.mean(sub_combined_utility_results['S->S']), width,
                   yerr=np.std(sub_combined_utility_results['S->S']),
                   color='b', alpha=alpha_value, capsize=error_bar_capsize)
            ax.bar(setting_idx + 1 + width * 2,
                   np.mean(sub_combined_utility_results['S->O']), width,
                   yerr=np.std(sub_combined_utility_results['S->O']),
                   color='y', alpha=alpha_value, capsize=error_bar_capsize)



        ax.set_ylabel('MSE')
        ax.set_xlabel('DP-GAN Scheme', labelpad=14)
        ax.set_xticks(np.arange(len(model_names) + 1) + width * 1)
        ax.set_xticklabels(['Original'] + [x.upper().strip('_64').replace('_', '-') for x in model_names])
        ax.set_ylim(0, 0.3)

        ax.legend(('Original model to original data (test MSE)',
                   'Original model to synthetic data (MSE: pred. labels vs syn. labels)',
                   'Synthetic model to synthetic data (test MSE)',
                   'Synthetic model to original data (MSE: pred. labels vs orig. labels)'),
                  loc='upper left')
        # plt.title('MSE Error of Synthetic Data Generation for varying model structure', fontdict={'size': 20})
        plt.savefig(f'mse_test_chen_custom_{eps_name}_week{week_name}.png')
        plt.close()
